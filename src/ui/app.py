"""Streamlit UI для чат-бота."""

import streamlit as st
import sys
import os
import uuid

import openpyxl
import io

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from src.agent.graph import create_graph
from src.agent.state import AgentState
from src.agent.schemas import AgentResponse, ErrorResponse

from src.config import callbacks

import logging
level_name = os.getenv("LOG_LEVEL", "INFO").upper()
level = getattr(logging, level_name, logging.INFO)
for h in logging.root.handlers[:]:
    logging.root.removeHandler(h)
    
logging.basicConfig(
level=level,
format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
handlers=[logging.StreamHandler(sys.stdout)],
)    
logging.getLogger().setLevel(level)
    

st.set_page_config(
    page_title="Ad Placement Agent",
    page_icon="📊",
    layout="wide",
)

st.title("📊 Агент планирования рекламы")
st.caption("Чат-бот для работы с планами размещения рекламы домашнего интернета")


# --- Sidebar: выбор пользователя ---
st.sidebar.header("Настройки")

USER_OPTIONS = {
    "editor_nsk_01": {"label": "Иванов И.И. (Editor, Новосибирск)", "role": "editor", "branch": "Новосибирск"},
    "editor_kzn_01": {"label": "Петров П.П. (Editor, Казань)", "role": "editor", "branch": "Казань"},
    "editor_msk_01": {"label": "Сидоров С.С. (Editor, Москва)", "role": "editor", "branch": "Москва"},
    "approver_01": {"label": "Щербаков С.А. (Approver, HQ)", "role": "approver", "branch": "HQ"},
}

selected_user = st.sidebar.selectbox(
    "Пользователь:",
    options=list(USER_OPTIONS.keys()),
    format_func=lambda x: USER_OPTIONS[x]["label"],
)

user_info = USER_OPTIONS[selected_user]

st.sidebar.divider()
st.sidebar.markdown(f"**Роль:** `{user_info['role']}`")
st.sidebar.markdown(f"**Филиал:** `{user_info['branch']}`")


# --- Chat history ---
if "messages" not in st.session_state:
    st.session_state.messages = []
    
if "graph" not in st.session_state:
    graph_obj = create_graph()
    try:
        dbg = graph_obj.get_graph() if hasattr(graph_obj, "get_graph") else None
        print("GRAPH NODES:", list(getattr(dbg, "nodes", [])))
        print("GRAPH EDGES:", list(getattr(dbg, "edges", [])))
    except Exception as e:
        print("GRAPH DEBUG ERROR:", e)
    st.session_state.graph = graph_obj

# Display chat history
for msg in st.session_state.messages:
    with st.chat_message(msg["role"]):
        st.markdown(msg["content"])


# --- File upload ---
uploaded_file = st.sidebar.file_uploader(
    "Загрузить файл корректировок (.xlsx)",
    type=["xlsx"],
    key="corrections_file",
)


# --- Chat input ---
if prompt := st.chat_input("Введите сообщение..."):
    # Show user message
    st.session_state.messages.append({"role": "user", "content": prompt})
    with st.chat_message("user"):
        st.markdown(prompt)

    # Parse file if attached
    corrections_file_content = None
    has_attachment = False
    file_path = None

    if uploaded_file is not None:

        has_attachment = True
        # Сохраняем файл для tool
        data = uploaded_file.read()
        os.makedirs("data/uploads", exist_ok=True)
        file_path = f"data/uploads/{uploaded_file.name}"
        with open(file_path, "wb") as f:
            f.write(data)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
        corrections_file_content = rows

    # Prepare state
    initial_state: AgentState = {
        # Пользователь
        "user_id": selected_user,
        "user_role": user_info["role"],
        "user_branch": user_info["branch"],
        "has_attachment": has_attachment,
        "file_path": file_path,
        "messages": [{"role": "user", "content": prompt}],

        # Роутинг
        "intent": None,
        "permission_granted": False,

        # Structured output
        "intent_data": None,
        "response": None,

        # План
        "target_month": None,
        "plan_exists": None,
        "plan_data": None,

        # Дедлайн
        "deadline_ok": None,
        "deadline_info": None,

        # Валидация корректировок
        "validation_passed": None,
        "validation_errors": None,
        "corrections_data": None,
        "corrections_file_content": corrections_file_content,

        # Approve / Finalize flow
        "approval_decision": None,              # "approve_branch" | "reject_branch" | "finalize_plan"
        "target_branch_for_approval": None,     # для совместимости с одиночным выбором
        "target_branches_for_approval": None,   # множественный выбор
        "rejection_reason": None,
        "approval_parse": None,                 # результат LLM-парсинга
        "plan_finalized": False,
        "plan_finalized_at": None,

        # Мета
        "request_id": str(uuid.uuid4()),
        "is_error": False,
        "iteration": 0,
    }

    # Run agent
    with st.chat_message("assistant"):
        with st.spinner("Думаю..."):
            try:
                print("DEBUG file_path:", file_path, "has_attachment:", has_attachment)
                result = st.session_state.graph.invoke(initial_state, config={"callbacks": callbacks})

                # Извлекаем structured response
                response_obj = result.get("response")
                file_path_from_state = result.get("file_path")

                if isinstance(response_obj, AgentResponse):
                    response_text = response_obj.message

                    # Показываем next_steps если есть
                    if response_obj.next_steps:
                        steps = "\n".join(f"- {s}" for s in response_obj.next_steps)
                        response_text += f"\n\nСледующие шаги:\n{steps}"

                elif isinstance(response_obj, ErrorResponse):
                    response_text = f"⚠️ {response_obj.message}"

                elif response_obj is None:
                    # Fallback: берём из messages (старый формат)
                    msgs = result.get("messages", [])
                    if msgs:
                        last_msg = msgs[-1]
                        response_text = last_msg.get("content", "Не удалось получить ответ.")
                    else:
                        response_text = "Не удалось получить ответ."
                else:
                    response_text = str(response_obj)

                st.markdown(response_text)
                st.session_state.messages.append(
                    {"role": "assistant", "content": response_text}
                )

                # Show file download if plan was exported
                if result.get("plan_exists") and result.get("plan_data"):
                    export_path = f"data/exports/plan_{user_info['branch']}_{result.get('target_month')}.xlsx"
                    if os.path.exists(export_path):
                        with open(export_path, "rb") as f:
                            st.download_button(
                                label="📥 Скачать план (Excel)",
                                data=f.read(),
                                file_name=f"plan_{result.get('target_month')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )

            except Exception as e:
                error_msg = f"⚠️ Ошибка: {str(e)}"
                st.error(error_msg)
                st.session_state.messages.append(
                    {"role": "assistant", "content": error_msg}
                )