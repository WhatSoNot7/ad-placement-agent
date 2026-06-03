"""Tool: экспорт плана в Excel."""

from pathlib import Path
from langchain_core.tools import tool

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@tool
def export_plan_to_excel(plan_data: list[dict], branch: str, month: str) -> dict:
    """
    Формирует Excel-файл из данных плана.

    Args:
        plan_data: Данные плана (список словарей)
        branch: Название филиала
        month: Месяц в формате YYYY-MM

    Returns:
        {"success": bool, "file_path": str, "message": str}
    """
    if not HAS_OPENPYXL:
        return {"success": False, "file_path": "", "message": "openpyxl не установлен"}

    if not plan_data:
        return {"success": False, "file_path": "", "message": "Нет данных для экспорта"}

    # Создаём workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"План {branch} {month}"

    # Заголовки
    headers = [
        "House ID",
        "Тип рекламы",
        "Частота",
        "Квартиры",
        "Действующие абоненты",
        "Прогноз заявок",
        "Стоимость",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    # Маппинг полей к колонкам
    field_map = [
        "house_id",
        "ad_type",
        "frequency",
        "apartments",
        "existing_subscribers",
        "predicted_leads",
        "cost",
    ]

    # Заполняем данные
    for row_idx, row_data in enumerate(plan_data, 2):
        for col_idx, field in enumerate(field_map, 1):
            value = row_data.get(field, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Строка итогов
    total_row = len(plan_data) + 2
    bold_font = Font(bold=True)

    total_label_cell = ws.cell(row=total_row, column=1, value="ИТОГО")
    total_label_cell.font = bold_font

    # Сумма прогноза заявок
    total_leads = 0
    for r in plan_data:
        leads_value = r.get("predicted_leads", 0)
        total_leads += leads_value

    total_leads_cell = ws.cell(row=total_row, column=6, value=total_leads)
    total_leads_cell.font = bold_font

    # Сумма стоимости
    total_cost = 0
    for r in plan_data:
        cost_value = r.get("cost", 0)
        total_cost += cost_value

    total_cost_cell = ws.cell(row=total_row, column=7, value=total_cost)
    total_cost_cell.font = bold_font

    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        column_letter = None
        for cell in col:
            if column_letter is None:
                column_letter = cell.column_letter
            try:
                cell_str = str(cell.value) if cell.value else ""
                cell_len = len(cell_str)
                if cell_len > max_length:
                    max_length = cell_len
            except TypeError:
                pass
        adjusted_width = min(max_length + 2, 30)
        if column_letter:
            ws.column_dimensions[column_letter].width = adjusted_width

    # Сохраняем файл
    output_dir = Path("data/exports")
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = f"plan_{branch}_{month}.xlsx"
    file_path = output_dir / filename
    wb.save(file_path)

    result = {
        "success": True,
        "file_path": str(file_path),
        "message": f"Файл сформирован: {filename}",
    }
    return result